from mail_func import Email
from openpyxl import load_workbook

wb = load_workbook('Email_list.xlsx')
sheet = wb.active

e = Email()
for row in sheet.iter_rows():
	name=row[0].value
	add=row[1].value
	e.send_mail(name,addr)